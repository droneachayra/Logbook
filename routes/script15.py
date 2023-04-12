from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, time
from openpyxl.styles import Font
import json
import csv
import math

import datetime as dt

# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
json_path = r"C:\Users\Admin\Desktop\Log Book Automation\Log Book Automation 12 trainees MiMiSm\3315.json"
excel_path = r'C:\Users\Admin\Desktop\Log Book Automation\Log Book Automation 12 trainees MiMiSm\3315.xlsx'
# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


with open (json_path) as file:
    data = json.load(file)

wb = Workbook()

header = ["Date", "UIN", "Class", "Category", "Trainer", "Trainee", "Place of Operation", "Start", "End", "Duration", "Dual", "RPIC", "CAT1", "EX"]

# day1 = data['DATE']
# day2 = day1 + 1
# day3 = day1 + 2

start_date = data['DATE']
start_date = dt.datetime.strptime(start_date, "%d-%m-%Y")

day1_temp = start_date
day1 = start_date.day
day2 = (day1_temp + dt.timedelta(days=1)).day
day3 = (day1_temp + dt.timedelta(days=2)).day
day4 = (day1_temp + dt.timedelta(days=3)).day

date1 = start_date.date()
date2 = (day1_temp + dt.timedelta(days=1)).date()
date3 = (day1_temp + dt.timedelta(days=2)).date()
date4 = (day1_temp + dt.timedelta(days=3)).date()



# clas = data['CLASS'][0]
category = data['Category'][0]
place = data["Place_of_Operation"]
trainees = data['Trainee']
count_of_trainers = len(data['Trainer'])
count_of_trainees = len(data['Trainee'])
count_of_drones = len(data['UIN'])
drone_index = 0
count_start_day1 = 0
count_start_day2 = 0
count = 0


#----------------------------------------------------------------------------
temprary = math.ceil(count_of_trainees / count_of_trainers)
#----------------------------------------------------------------------------


ex_day1 = ["EX-01", "EX-02", "EX-03", "EX-04", "EX-05", "EX-06", "EX-07"]
ex_day2 = ["EX-08A", "EX-08B", "PROGRESS CHECK", "EX-10", "EX-11A", "EX-11B", "EX-12A", "EX-12B", "EX-13", "EX-14", "EX-15A", "EX-15B"]
drone_log_header = ["SL No.", "Date", "Name of RPIC", "Place of Operation", "Start Time", "End Time", "Duration", "Cumulative Hrs", "Remarks"]

ws1 = wb.active
ws1.title = "DAY1"
ws1 = wb["DAY1"]

wb.create_sheet(index = 1, title = 'DAY2')
ws2 = wb['DAY2']

wb.create_sheet(index = 2, title = 'DAY3')
ws3 = wb['DAY3']

wb.create_sheet(index = 3, title = 'EXAM')
ws4 = wb['EXAM']


drone1 = data['UIN'][0]
wb.create_sheet(index = 4, title = drone1)
ws5 = wb[drone1]
ws5.append(drone_log_header)
drone1_day1 = []
drone1_day2 = []

if(len(data["UIN"]) > 1):
    drone2 = data["UIN"][1]
    wb.create_sheet(index = 5, title = drone2)
    ws6 = wb[drone2]
    ws6.append(drone_log_header)
    drone2_day1 = []
    drone2_day2 = []

if(len(data["UIN"]) == 3):
    drone3 = data["UIN"][2]
    wb.create_sheet(index = 6, title = drone3)
    ws7 = wb[drone3]
    ws7.append(drone_log_header)
    drone3_day1 = []
    drone3_day2 = []
    


# drone2_day1 = []
# drone2_day2 = []


slno = cumulative_Hrs = remarks = "" 
temp_count = 0
# ++++++++++++++++++++++++++++++++++++ DAY 1 Excersice 1 ++++++++++++++++++++++++++++++++++++++++++++++++++++ 

for student in range(9):

    if(temp_count == 3):
        drone_index += 1
        temp_count = 0

    if(count_start_day1 == 3):
        count_start_day1 = 0
        st = data['Start_day1'][count_start_day1]
        x = st.split(":")
        h = int(x[0])
        m = int(x[1])
        total_time = datetime(1, 1, 1, 0, 0, 0)
        st = datetime(1, 1, 1, h, m, 0)
        et = datetime(1, 1, 1, 0, 0, 0)
        count_start_day1 += 1
        count += 1

    else:
        st = data['Start_day1'][count_start_day1]
        x = st.split(":")
        h = int(x[0])
        m = int(x[1])
        total_time = datetime(1, 1, 1, 0, 0, 0)
        st = datetime(1, 1, 1, h, m, 0)
        et = datetime(1, 1, 1, 0, 0, 0)
        count_start_day1 += 1

    trainee = trainees[student]

    temp = []
    ws1.append(temp)
    ws1.append(header)

    uin = data['UIN'][drone_index]
    if(uin == 'UA000GF'):
        clas = 'Sm'
    else:
        clas = 'Mi'

    
    for y in range(len(ex_day1)):
        flag1 = 0  
        values = []
        trer = data['Trainer'][count]
        ex = ex_day1[y]
        
        if(ex == 'EX-01'):
            et = st + timedelta(minutes=30)
            total_time += timedelta(minutes=30)
            flag1 = 1
            duration = dual = cat1 = "00:30"

        else:
            et = st + timedelta(minutes=15)
            total_time += timedelta(minutes=15)
            duration = dual = cat1 = "00:15"
        
        if(total_time == datetime(1, 1, 1, 1, 30)):
            total_time = datetime(1, 1, 1, 0, 0, 0)
            st += timedelta(minutes=15)
            et += timedelta(minutes=15)
                    
        #------------------------------------------------------
        start = str(st.time())
        start = start.split(":")
        start = start[0] + ":" + start[1]


        end = str(et.time())
        end = end.split(":")
        end = end[0] + ":" + end[1]
        #------------------------------------------------------

        rpic = ""
        ws1.append([day1, uin, clas, category, trer, trainee, place, start, end, duration, dual, rpic, cat1, ex])
        # print(">>>>>>", [day1, uin, clas, category, trer, trainee, place, start, end, duration, dual, rpic, cat1, ex])
        # print(start, end)
                
        if(uin == drone1):
            # ws4.append([slno, day1, trer, place, start, end, duration, cumulative_Hrs, remarks])
            drone1_day1.append([slno, date1, trer, place, start, end, duration, cumulative_Hrs, remarks])
        
        elif(uin == drone2):
            # ws5.append([slno, day1, trer, place, start, end, duration, cumulative_Hrs, remarks])
            drone2_day1.append([slno, date1, trer, place, start, end, duration, cumulative_Hrs, remarks])
        
        else:
            # ws5.append([slno, day1, trer, place, start, end, duration, cumulative_Hrs, remarks])
            drone3_day1.append([slno, date1, trer, place, start, end, duration, cumulative_Hrs, remarks])


        if(flag1 == 1):
            st += timedelta(minutes=30)
            flag1 = 0
        else:
            st+= timedelta(minutes=15)
    
    temp_count += 1



# clas = data['CLASS'][0]
category = data['Category'][0]
place = data["Place_of_Operation"]
trainees = data['Trainee']
count_of_trainers = len(data['Trainer'])
count_of_trainees = len(data['Trainee'])
count_of_drones = len(data['UIN'])
drone_index = 0
count_start_day1 = 0
count_start_day2 = 0
count = 0
temp_count = 0

for student in range(9, 14):

    if(temp_count == 3):
        drone_index += 1
        temp_count = 0

    if(count_start_day1 == 3):
        count_start_day1 = 0
        st = data['Start_day1'][count_start_day1]
        x = st.split(":")
        h = int(x[0])
        m = int(x[1])
        total_time = datetime(1, 1, 1, 0, 0, 0)
        st = datetime(1, 1, 1, h, m, 0)
        et = datetime(1, 1, 1, 0, 0, 0)
        count_start_day1 += 1
        count += 1

    else:
        st = data['Start_day1'][count_start_day1]
        x = st.split(":")
        h = int(x[0])
        m = int(x[1])
        total_time = datetime(1, 1, 1, 0, 0, 0)
        st = datetime(1, 1, 1, h, m, 0)
        et = datetime(1, 1, 1, 0, 0, 0)
        count_start_day1 += 1

    trainee = trainees[student]

    temp = []
    ws2.append(temp)
    ws2.append(header)

    uin = data['UIN'][drone_index]
    if(uin == 'UA000GF'):
        clas = 'Sm'
    else:
        clas = 'Mi'

    
    for y in range(len(ex_day1)):
        flag1 = 0  
        values = []
        trer = data['Trainer'][count]
        ex = ex_day1[y]
        
        if(ex == 'EX-01'):
            et = st + timedelta(minutes=30)
            total_time += timedelta(minutes=30)
            flag1 = 1
            duration = dual = cat1 = "00:30"

        else:
            et = st + timedelta(minutes=15)
            total_time += timedelta(minutes=15)
            duration = dual = cat1 = "00:15"
        
        if(total_time == datetime(1, 1, 1, 1, 30)):
            total_time = datetime(1, 1, 1, 0, 0, 0)
            st += timedelta(minutes=15)
            et += timedelta(minutes=15)
                    
        #------------------------------------------------------
        start = str(st.time())
        start = start.split(":")
        start = start[0] + ":" + start[1]


        end = str(et.time())
        end = end.split(":")
        end = end[0] + ":" + end[1]
        #------------------------------------------------------

        rpic = ""
        ws2.append([day2, uin, clas, category, trer, trainee, place, start, end, duration, dual, rpic, cat1, ex])
        # print(">>>>>>", [day1, uin, clas, category, trer, trainee, place, start, end, duration, dual, rpic, cat1, ex])
        # print(start, end)
                
        if(uin == drone1):
            # ws4.append([slno, day1, trer, place, start, end, duration, cumulative_Hrs, remarks])
            drone1_day2.append([slno, date2, trer, place, start, end, duration, cumulative_Hrs, remarks])
        
        elif(uin == drone2):
            # ws5.append([slno, day1, trer, place, start, end, duration, cumulative_Hrs, remarks])
            drone2_day2.append([slno, date2, trer, place, start, end, duration, cumulative_Hrs, remarks])
        
        else:
            # ws5.append([slno, day1, trer, place, start, end, duration, cumulative_Hrs, remarks])
            drone3_day2.append([slno, date2, trer, place, start, end, duration, cumulative_Hrs, remarks])


        if(flag1 == 1):
            st += timedelta(minutes=30)
            flag1 = 0
        else:
            st+= timedelta(minutes=15)

    temp_count += 1

print("Hold")
print("Hold here")
# # +++++++++++++++++++++++++++++++++++ DAY 2 +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

# category = data['Category'][0]
# place = data["Place_of_Operation"]
# trainees = data['Trainee']
# count_of_trainers = len(data['Trainer'])
# count_of_trainees = len(data['Trainee'])
# count_of_drones = len(data['UIN'])
drone_index = 1
# count_start_day1 = 0
# count_start_day2 = 0
count = 1
temp_count = 0


count_start_day2 = 2
for student in range(0, 4):
    flag2 = 0
    # if(temp_count == 3):
    #     drone_index += 1
    #     temp_count = 0
    

    if(count_start_day2 == 3):
        count_start_day2 = 0
        st = data['Start_day2'][count_start_day2]
        x = st.split(":")
        h = int(x[0])
        m = int(x[1])
        total_time1 = datetime(1, 1, 1, 0, 0, 0)
        st = datetime(1, 1, 1, h, m, 0)
        et = datetime(1, 1, 1, 0, 0, 0)
        count_start_day2 += 1
        count += 1

    else:
        st = data['Start_day2'][count_start_day2]
        x = st.split(":")
        h = int(x[0])
        m = int(x[1])
        total_time1 = datetime(1, 1, 1, 0, 0, 0)
        st = datetime(1, 1, 1, h, m, 0)
        et = datetime(1, 1, 1, 0, 0, 0)
        count_start_day2 += 1

    trainee = trainees[student]

    temp = []
    ws2.append(temp)
    ws2.append(header)


    for y in range(len(ex_day2)):  
        values = []
        trer = data['Trainer'][count]
        ex = ex_day2[y]
        
        if('A' in ex):
            et = st + timedelta(minutes=5)
            total_time1 += timedelta(minutes=5)
            flag2 = 1
            duration = dual = cat1 = "00:05"
            rpic = ""

        elif('B' in ex):
            et = st + timedelta(minutes=10)
            total_time1 += timedelta(minutes=10)
            duration = rpic = cat1 = "00:10"
            dual = ""
            flag2 = 2
        
        else:
            et = st + timedelta(minutes=15)
            total_time1 += timedelta(minutes=15)
            if(ex == "PROGRESS CHECK"):
                duration = rpic = cat1 = "00:15"
                dual = ""
            else:
                duration = dual = cat1 = "00:15"
                rpic = ""
        
        if(total_time1 == datetime(1, 1, 1, 1, 15)):
            print("*************")
            # print(total_time1)
            st += timedelta(minutes=15)
            et += timedelta(minutes=15)
            total_time1 = datetime(1, 1, 1, 0, 0, 0)
            
        uin = data['UIN'][drone_index]    
        #------------------------------------------------------
        start1 = str(st.time())
        start1 = start1.split(":")
        start1 = start1[0] + ":" + start1[1]
        # print(start1)

        end1 = str(et.time())
        end1 = end1.split(":")
        end1 = end1[0] + ":" + end1[1]
        # print(end1)
        #------------------------------------------------------

        ws2.append([day2, uin, clas, category, trer, trainee, place, start1, end1, duration, dual, rpic, cat1, ex])
        # ws2.append([day2, uin, clas, category, trer, trainee, place, st.time(), et.time(), duration, dual, rpic, cat1, ex])
        print([day2, uin, clas, category, trer, trainee, place, start1, end1, duration, dual, rpic, cat1, ex])

        test_header = [day2, uin, clas, category, trer, trainee, place, start1, end1, duration, dual, rpic, cat1, ex]
        # print("ST", st, "et", et)

        
        if(uin == drone1):
            if(rpic == ""):
                # ws4.append([slno, day2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone1_day2.append([slno, date2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])

            else:
                # ws4.append([slno, day2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone1_day2.append([slno, date2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
        
        elif(uin == drone2):
            if(rpic == ""):
                # ws5.append([slno, day2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone2_day2.append([slno, date2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
            else:
                # ws5.append([slno, day2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone2_day2.append([slno, date2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])

        else:
            if(rpic == ""):
                # ws5.append([slno, day2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone3_day2.append([slno, date2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
            else:
                # ws5.append([slno, day2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone3_day2.append([slno, date2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
        
        if(flag2 == 1):
            st += timedelta(minutes=5)
            flag2 = 0

        elif(flag2 == 2):
            st += timedelta(minutes=10)
            flag2 = 0

        else:
            st+= timedelta(minutes=15)
    temp_count += 1
    if(temp_count == 1):
        drone_index += 1


# # ++++++++++++++++++++++++++++++++ END of DAY 2 ++++++++++++++++++++++++++++++++++++++++++++++++++++ 

# # ++++++++++++++++++++++++++++++++++++ DAY 3 ++++++++++++++++++++++++++++++++++++++++++++++++++++ 
drone_index = 0
# count_start_day1 = 0
# count_start_day2 = 0
count = 0
temp_count = 0


count_start_day2 = 0
for student in range(5, 14):
    flag2 = 0
    if(temp_count == 3):
        drone_index += 1
        temp_count = 0
    

    if(count_start_day2 == 3):
        count_start_day2 = 0
        st = data['Start_day2'][count_start_day2]
        x = st.split(":")
        h = int(x[0])
        m = int(x[1])
        total_time1 = datetime(1, 1, 1, 0, 0, 0)
        st = datetime(1, 1, 1, h, m, 0)
        et = datetime(1, 1, 1, 0, 0, 0)
        count_start_day2 += 1
        count += 1

    else:
        st = data['Start_day2'][count_start_day2]
        x = st.split(":")
        h = int(x[0])
        m = int(x[1])
        total_time1 = datetime(1, 1, 1, 0, 0, 0)
        st = datetime(1, 1, 1, h, m, 0)
        et = datetime(1, 1, 1, 0, 0, 0)
        count_start_day2 += 1

    trainee = trainees[student]

    temp = []
    ws3.append(temp)
    ws3.append(header)


    for y in range(len(ex_day2)):  
        values = []
        trer = data['Trainer'][count]
        ex = ex_day2[y]
        
        if('A' in ex):
            et = st + timedelta(minutes=5)
            total_time1 += timedelta(minutes=5)
            flag2 = 1
            duration = dual = cat1 = "00:05"
            rpic = ""

        elif('B' in ex):
            et = st + timedelta(minutes=10)
            total_time1 += timedelta(minutes=10)
            duration = rpic = cat1 = "00:10"
            dual = ""
            flag2 = 2
        
        else:
            et = st + timedelta(minutes=15)
            total_time1 += timedelta(minutes=15)
            if(ex == "PROGRESS CHECK"):
                duration = rpic = cat1 = "00:15"
                dual = ""
            else:
                duration = dual = cat1 = "00:15"
                rpic = ""
        
        if(total_time1 == datetime(1, 1, 1, 1, 15)):
            print("*************")
            # print(total_time1)
            st += timedelta(minutes=15)
            et += timedelta(minutes=15)
            total_time1 = datetime(1, 1, 1, 0, 0, 0)
            
        uin = data['UIN'][drone_index]    
        #------------------------------------------------------
        start1 = str(st.time())
        start1 = start1.split(":")
        start1 = start1[0] + ":" + start1[1]
        # print(start1)

        end1 = str(et.time())
        end1 = end1.split(":")
        end1 = end1[0] + ":" + end1[1]
        # print(end1)
        #------------------------------------------------------

        ws3.append([day3, uin, clas, category, trer, trainee, place, start1, end1, duration, dual, rpic, cat1, ex])
        # ws2.append([day2, uin, clas, category, trer, trainee, place, st.time(), et.time(), duration, dual, rpic, cat1, ex])
        print([day3, uin, clas, category, trer, trainee, place, start1, end1, duration, dual, rpic, cat1, ex])

        test_header = [day3, uin, clas, category, trer, trainee, place, start1, end1, duration, dual, rpic, cat1, ex]
        # print("ST", st, "et", et)

        
        if(uin == drone1):
            if(rpic == ""):
                # ws4.append([slno, day2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone1_day2.append([slno, date3, trer, place, start1, end1, duration, cumulative_Hrs, remarks])

            else:
                # ws4.append([slno, day2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone1_day2.append([slno, date3, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
        
        elif(uin == drone2):
            if(rpic == ""):
                # ws5.append([slno, day2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone2_day2.append([slno, date3, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
            else:
                # ws5.append([slno, day2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone2_day2.append([slno, date3, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])

        else:
            if(rpic == ""):
                # ws5.append([slno, day2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone3_day2.append([slno, date3, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
            else:
                # ws5.append([slno, day2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
                drone3_day2.append([slno, date3, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
        
        if(flag2 == 1):
            st += timedelta(minutes=5)
            flag2 = 0

        elif(flag2 == 2):
            st += timedelta(minutes=10)
            flag2 = 0

        else:
            st+= timedelta(minutes=15)
    temp_count += 1
    # if(temp_count == 2):
    #     drone_index += 1



drone_index = 0
# count_start_day1 = 0
# count_start_day2 = 0
count = 0
temp_count = 0

count_start_day2 = 0


if(count_start_day2 == 3):
    count_start_day2 = 0
    st = data['Start_day2'][count_start_day2]
    x = st.split(":")
    h = int(x[0])
    m = int(x[1])
    total_time1 = datetime(1, 1, 1, 0, 0, 0)
    st = datetime(1, 1, 1, h, m, 0)
    et = datetime(1, 1, 1, 0, 0, 0)
    count_start_day2 += 1
    count += 1

else:
    st = data['Start_day2'][count_start_day2]
    x = st.split(":")
    h = int(x[0])
    m = int(x[1])
    total_time1 = datetime(1, 1, 1, 0, 0, 0)
    st = datetime(1, 1, 1, h, m, 0)
    et = datetime(1, 1, 1, 0, 0, 0)
    count_start_day2 += 1
student = 14
trainee = trainees[student]

temp = []
ws4.append(temp)
ws4.append(header)


for y in range(len(ex_day2)):  
    values = []
    trer = data['Trainer'][count]
    ex = ex_day2[y]
    
    if('A' in ex):
        et = st + timedelta(minutes=5)
        total_time1 += timedelta(minutes=5)
        flag2 = 1
        duration = dual = cat1 = "00:05"
        rpic = ""

    elif('B' in ex):
        et = st + timedelta(minutes=10)
        total_time1 += timedelta(minutes=10)
        duration = rpic = cat1 = "00:10"
        dual = ""
        flag2 = 2
    
    else:
        et = st + timedelta(minutes=15)
        total_time1 += timedelta(minutes=15)
        if(ex == "PROGRESS CHECK"):
            duration = rpic = cat1 = "00:15"
            dual = ""
        else:
            duration = dual = cat1 = "00:15"
            rpic = ""
    
    if(total_time1 == datetime(1, 1, 1, 1, 15)):
        print("*************")
        # print(total_time1)
        st += timedelta(minutes=15)
        et += timedelta(minutes=15)
        total_time1 = datetime(1, 1, 1, 0, 0, 0)
        
    uin = data['UIN'][drone_index]    
    #------------------------------------------------------
    start1 = str(st.time())
    start1 = start1.split(":")
    start1 = start1[0] + ":" + start1[1]
    # print(start1)

    end1 = str(et.time())
    end1 = end1.split(":")
    end1 = end1[0] + ":" + end1[1]
    # print(end1)
    #------------------------------------------------------

    ws4.append([day4, uin, clas, category, trer, trainee, place, start1, end1, duration, dual, rpic, cat1, ex])
    # ws2.append([day2, uin, clas, category, trer, trainee, place, st.time(), et.time(), duration, dual, rpic, cat1, ex])
    print([day4, uin, clas, category, trer, trainee, place, start1, end1, duration, dual, rpic, cat1, ex])

    test_header = [day4, uin, clas, category, trer, trainee, place, start1, end1, duration, dual, rpic, cat1, ex]
    # print("ST", st, "et", et)

    
    if(uin == drone1):
        if(rpic == ""):
            # ws4.append([slno, day2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
            drone1_day2.append([slno, date4, trer, place, start1, end1, duration, cumulative_Hrs, remarks])

        else:
            # ws4.append([slno, day2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
            drone1_day2.append([slno, date4, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
    
    elif(uin == drone2):
        if(rpic == ""):
            # ws5.append([slno, day2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
            drone2_day2.append([slno, date4, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
        else:
            # ws5.append([slno, day2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
            drone2_day2.append([slno, date4, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])

    else:
        if(rpic == ""):
            # ws5.append([slno, day2, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
            drone3_day2.append([slno, date4, trer, place, start1, end1, duration, cumulative_Hrs, remarks])
        else:
            # ws5.append([slno, day2, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
            drone3_day2.append([slno, date4, trainee, place, start1, end1, duration, cumulative_Hrs, remarks])
    
    if(flag2 == 1):
        st += timedelta(minutes=5)
        flag2 = 0

    elif(flag2 == 2):
        st += timedelta(minutes=10)
        flag2 = 0

    else:
        st+= timedelta(minutes=15)
temp_count += 1
# if(temp_count == 2):
#     drone_index += 1

ws4.append(temp)



# # +++++++++++++++++++++++++++++++++++++++++++++ EXAM DAY ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
examiner = data["EXAMINER"][0]
exam_start = data["EXAM_TIME"]
exam_uin = data["EXAM_UIN"][0]
ex = "FINAL TEST"

st = exam_start[0]
x = st.split(":")
h = int(x[0])
m = int(x[1])
# total_time = datetime(1, 1, 1, 0, 0, 0)
st = datetime(1, 1, 1, h, m, 0)
et = datetime(1, 1, 1, 0, 0, 0)

duration = rpic = cat1 = "15:00"
dual = ""
for student in range(len(trainees)):
    ws4.append(header)
    

    et = st + timedelta(minutes=15)

    if(student == 6):
        st =  st + timedelta(minutes=30)
        et = st + timedelta(minutes=30)


    start = str(st.time())
    start = start.split(":")
    start = start[0] + ":" + start[1]

    end = str(et.time())
    end = end.split(":")
    end = end[0] + ":" + end[1]

    values = []
    trainee = trainees[student]
    ws4.append([day4, exam_uin, clas, category, examiner, trainee, place, start, end, duration, dual, rpic, cat1, ex])
    ws4.append(values)

    try:
        if(drone1 == exam_uin):
            # ws5.append([slno, day3, trainee, place, start, end, duration, cumulative_Hrs, remarks])
            drone1_day2.append([slno, date4, trainee, place, start, end, duration, cumulative_Hrs, remarks])
        elif(drone2 == exam_uin):
            drone2_day2.append([slno, date4, trainee, place, start, end, duration, cumulative_Hrs, remarks])
        else:
            drone3_day2.append([slno, date4, trainee, place, start, end, duration, cumulative_Hrs, remarks])
    
    except:
        pass
    st += timedelta(minutes=30)



for x in range(len(drone1_day1)):
    ws5.append(drone1_day1[x])

for x in range(len(drone1_day2)):
    ws5.append(drone1_day2[x])

if(len(data["UIN"]) > 1):
    for x in range(len(drone2_day1)):
        ws6.append(drone2_day1[x])

    for x in range(len(drone2_day2)):
        ws6.append(drone2_day2[x])

if(len(data["UIN"]) == 3):
    for x in range(len(drone3_day1)):
        ws7.append(drone3_day1[x])

    for x in range(len(drone3_day2)):
        ws7.append(drone3_day2[x])

# ws5.append(drone1_day1)
# ws4.append(drone1_day2)
wb.save(excel_path)


